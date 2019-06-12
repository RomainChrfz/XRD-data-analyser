import sys, os, platform
from scipy.optimize import curve_fit
import openpyxl as op
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import Cell
import pdb
import numpy as np
from matplotlib.backends.backend_qt5agg import  \
    FigureCanvasQTAgg as FigureCanvas,          \
    NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from PyQt5.QtWidgets import (QApplication, QMainWindow, QDesktopWidget,
                             QTabWidget, QAction, QMessageBox, QTextEdit,
                             QFileDialog, QWidget, QPushButton, QVBoxLayout,
                             QHBoxLayout, QComboBox, QRubberBand, QLabel, QFrame,
                             QGridLayout, QLineEdit, QGroupBox, QProgressBar,
                             QToolTip, QCheckBox, QSlider, QComboBox, QMenu)
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import QCoreApplication, QPoint, Qt, QMargins

def resource_path(relative_path) :
    """Get absolute path to resource, works for dev and for PyInstaller"""

    try :
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except :
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class DataAnalyser(QMainWindow):

    icone_dir   = "icones"          # Icon directory
    cur_dir     = os.getcwd()+"/"   # Work directory 

    def __init__(self):

        super().__init__()
        # Attributs
        self.menubar = self.menuBar() # Needed for Mac

        # To manage tabs :
        self.tabs = QTabWidget()
        
        # Several attributs
        self.clplt = True
        
        # Tabs of the soft :        
        self.onePlotTab  = OnePlot(self)

        self.__initUI()   # Initialization of the user interface
        self.show()       # Display

    def center(self):
        '''To center the window on the screen'''
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)

    def __initUI(self):
        '''Creation of graphical elements'''
        self.resize(1280,720)
        self.center()
        self.setWindowTitle('Data Analyser') # Name of the window
        self.statusBar()  # Activation of the status bar

        self.tabs.addTab(self.onePlotTab,"Graph") # Name of the tab

        self.setCentralWidget(self.tabs)

        if platform.uname().system.startswith('Darw') :
            # Needed if the soft is launched on an Apple device :
            self.menubar.setNativeMenuBar(False)

        ###### Several Menu
        fileMenu = self.menubar.addMenu('File')
        optionMenu = self.menubar.addMenu('Options')
        creditMenu = self.menubar.addMenu('?')
        
        qa = QAction('Open XRD data File',self)
        qa.setShortcut('Ctrl+O')
        qa.setStatusTip('''Open a (.dql) XRD data file''')
        qa.triggered.connect(self.onePlotTab.open_file)
        fileMenu.addAction(qa)

        qa = QAction('Open text data File',self)
        qa.setShortcut('Ctrl+T')
        qa.setStatusTip('''Open a (.txt) text data file''')
        qa.triggered.connect(self.onePlotTab.open_text_file)
        fileMenu.addAction(qa)

        qa = QAction('Open excel data File',self)
        qa.setShortcut('Ctrl+F')
        qa.setStatusTip('''Open a (.xslx) excel data file''')
        qa.triggered.connect(self.onePlotTab.open_excel_file)
        fileMenu.addAction(qa)

        qa = QAction('Exit',self)
        qa.setShortcut('Ctrl+Q')
        qa.setStatusTip('''Close the software''')
        qa.triggered.connect(self.close)
        fileMenu.addAction(qa)

        qa = QAction('Display Status Bar',self,checkable = True)
        qa.setStatusTip('''To choose whether or not the status bar is displayed''')
        qa.setChecked(True)
        qa.triggered.connect(self.toggleSBar)
        optionMenu.addAction(qa)
        self.statBar = self.statusBar()


        qa = QAction('Erase graph before new',self,checkable = True)
        qa.setStatusTip('''Automatically erase the graph before new data file''')
        qa.setChecked(True)
        qa.triggered.connect(self.clearnewPlot)
        optionMenu.addAction(qa)

        qa = QAction('Information',self)
        qa.setStatusTip('If you need information on the soft')
        qa.triggered.connect(self.help_info)
        creditMenu.addAction(qa)
        
    def closeEvent(self, event):        
        rep = QMessageBox.question(self, 'Caution', "Do you really want to exit ?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)                  
        if rep == QMessageBox.Yes :
            event.accept()  
        else:
            event.ignore()   

    def clearnewPlot(self,state) :
        if state :
            self.clplt = True
        else :
            self.clplt = False
        
    def toggleSBar(self,state) :
        if state:
            self.statBar.show()
        else:
            self.statBar.hide()

    def clearPlots(self): 
        self.onePlotTab.ClearAxes()

    def help_info(self) :
        q = QMessageBox(self)
        q.setWindowTitle('Help')
        q.setText('<b>Software information<\b>')
        q.setInformativeText('* Version : 1.0\n* Author : Romain Charfaz\n* Contact : charfaz-pro@outlook.fr\n\nThis soft automatically plots the curve contained in your measurment file. It automatically detects peaks (amplitude > 20) up to a maximum of 10 different peaks (for now at least) and automatically fit a gaussian onto it. Therefore, based on each gaussian found, it automatically find the center, amplitude and full width half maximum and plot it next to the graph.')
        q.exec()


class OnePlot(QWidget):
    ''' Graph Widget '''

    def resource_path(relative_path) :
        """Get absolute path to resource, works for dev and for PyInstaller"""

        try :
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except :
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self, mainWindow):

        QWidget.__init__(self, mainWindow)

        self.mw = mainWindow

        self.figure    = None  
        self.axes      = None  
        self.canvas    = None  
        self.toolbar   = None  
        self.plot_xlim = None  
        self.plot_ylim = None  

        #Various attributs
        self.file_path = None
        self.x = None
        self.y = None
        self.line_num = None
        self.excel_line_num = None
        self.excel_column_num = None
        self.excel_sheet_name = None
        self.check = True
        self.sepa = None
        self.pic_num = None

        #Logo EPFL
        self.img_lbl = QLabel(self)
        pixmap = QPixmap(resource_path('./icones/epfl_logo.png'))
        self.img_lbl.setPixmap(pixmap)

        #Logo ANEMS
        #self.img_lbl2 = QLabel(self)
        #pixmap = QPixmap(resource_path('./icones/anems_logo.png'))
        #self.img_lbl2.setPixmap(pixmap)

        #Combo Box
        self.methode = QComboBox(self)
        self.methode.addItems(['...','Gaussian Fit','Moving Average','Linear Fit'])
        self.methode.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.methode.activated.connect(self.combobox)
        self.methode_label = QLabel('Method :')
        self.barre = QLabel('|')
        self.barre.hide()

        #Moving average
        self.mov_avg_btn = QPushButton('Moving Average',self)
        self.mov_avg_btn.clicked.connect(self.moving_avrg)
        self.mov_avg_btn.setEnabled(False)
        self.mov_avg_btn.setStatusTip('Apply a moving average to the data')
        self.mov_avg_val = QLineEdit()
        self.mov_avg_val.textEdited[str].connect(self.mov_avg_value)
        self.mov_avg_val.setMaxLength(3)
        self.mov_avg_val.setMaximumWidth(40)
        self.mov_avg_val.setEnabled(False)
        self.mov_avg_val.setStatusTip('Number of values on both sides of the current value taken into account for the computation of the mean')
        self.mov_avg_num = None
        self.mov_avg_btn.hide()
        self.mov_avg_val.hide()

        #Gaussian fit
        self.pic_1 = QLabel(self)
        self.threshold = 70
        self.gaus_btn = QPushButton('Gaussian Fit',self)
        self.gaus_btn.clicked.connect(self.gaussian_fit)
        self.gaus_btn.setEnabled(False)
        self.gaus_btn.setStatusTip('Apply a Gaussain Fit to the data')
        self.rock_check = QCheckBox('Rocking Curve ?')
        self.rock_check.setChecked(False)
        self.rock_check.stateChanged.connect(self.rocking_curve_active)
        self.test_rock = False
        self.gaus_btn.hide()
        self.rock_check.hide()

        #Linear fit
        self.lin_btn = QPushButton('Linear Fit',self)
        self.lin_btn.clicked.connect(self.linear_fit)
        self.lin_btn.setEnabled(False)
        self.lin_btn.setStatusTip('Apply a linear regression to fit the data')
        self.lin_btn.hide()

        self.__initUI()   # Initialization of the user interface

    def __initUI(self):

        vbox = QVBoxLayout()
        self.setLayout(vbox)

        #Info boxes
        self.gb1 = QGroupBox('Peak n°1')
        self.gb2 = QGroupBox('Peak n°2')
        self.gb3 = QGroupBox('Peak n°3')
        self.gb4 = QGroupBox('Peak n°4')
        self.gb5 = QGroupBox('Peak n°5')
        self.gb6 = QGroupBox('Peak n°6')
        self.gb7 = QGroupBox('Peak n°7')
        self.gb8 = QGroupBox('Peak n°8')
        self.gb9 = QGroupBox('Peak n°9')
        self.gb10 = QGroupBox('Peak n°10')
        self.amp_1 = QLabel(self)
        self.amp_2 = QLabel(self)
        self.amp_3 = QLabel(self)
        self.amp_4 = QLabel(self)
        self.amp_5 = QLabel(self)
        self.amp_6 = QLabel(self)
        self.amp_7 = QLabel(self)
        self.amp_8 = QLabel(self)
        self.amp_9 = QLabel(self)
        self.amp_10 = QLabel(self)
        self.center_1 = QLabel(self)
        self.center_2 = QLabel(self)
        self.center_3 = QLabel(self)
        self.center_4 = QLabel(self)
        self.center_5 = QLabel(self)
        self.center_6 = QLabel(self)
        self.center_7 = QLabel(self)
        self.center_8 = QLabel(self)
        self.center_9 = QLabel(self)
        self.center_10 = QLabel(self)
        self.fwhm_1 = QLabel(self)
        self.fwhm_2 = QLabel(self)
        self.fwhm_3 = QLabel(self)
        self.fwhm_4 = QLabel(self)
        self.fwhm_5 = QLabel(self)
        self.fwhm_6 = QLabel(self)
        self.fwhm_7 = QLabel(self)
        self.fwhm_8 = QLabel(self)
        self.fwhm_9 = QLabel(self)
        self.fwhm_10 = QLabel(self)
        vbox_1 = QVBoxLayout()
        vbox_2 = QVBoxLayout()
        vbox_3 = QVBoxLayout()
        vbox_4 = QVBoxLayout()
        vbox_5 = QVBoxLayout()
        vbox_6 = QVBoxLayout()
        vbox_7 = QVBoxLayout()
        vbox_8 = QVBoxLayout()
        vbox_9 = QVBoxLayout()
        vbox_10 = QVBoxLayout()
        vbox_1.addWidget(self.amp_1)
        vbox_1.addWidget(self.center_1)
        vbox_1.addWidget(self.fwhm_1)
        vbox_2.addWidget(self.amp_2)
        vbox_2.addWidget(self.center_2)
        vbox_2.addWidget(self.fwhm_2)
        vbox_3.addWidget(self.amp_3)
        vbox_3.addWidget(self.center_3)
        vbox_3.addWidget(self.fwhm_3)
        vbox_4.addWidget(self.amp_4)
        vbox_4.addWidget(self.center_4)
        vbox_4.addWidget(self.fwhm_4)
        vbox_5.addWidget(self.amp_5)
        vbox_5.addWidget(self.center_5)
        vbox_5.addWidget(self.fwhm_5)
        vbox_6.addWidget(self.amp_6)
        vbox_6.addWidget(self.center_6)
        vbox_6.addWidget(self.fwhm_6)
        vbox_7.addWidget(self.amp_7)
        vbox_7.addWidget(self.center_7)
        vbox_7.addWidget(self.fwhm_7)
        vbox_8.addWidget(self.amp_8)
        vbox_8.addWidget(self.center_8)
        vbox_8.addWidget(self.fwhm_8)
        vbox_9.addWidget(self.amp_9)
        vbox_9.addWidget(self.center_9)
        vbox_9.addWidget(self.fwhm_9)
        vbox_10.addWidget(self.amp_10)
        vbox_10.addWidget(self.center_10)
        vbox_10.addWidget(self.fwhm_10)
        self.gb1.setLayout(vbox_1)
        self.gb2.setLayout(vbox_2)
        self.gb3.setLayout(vbox_3)
        self.gb4.setLayout(vbox_4)
        self.gb5.setLayout(vbox_5)
        self.gb6.setLayout(vbox_6)
        self.gb6.setLayout(vbox_7)
        self.gb6.setLayout(vbox_8)
        self.gb6.setLayout(vbox_9)
        self.gb6.setLayout(vbox_10)
        self.gb1.hide()
        self.gb2.hide()
        self.gb3.hide()
        self.gb4.hide()
        self.gb5.hide()
        self.gb6.hide()
        self.gb7.hide()
        self.gb8.hide()
        self.gb9.hide()
        self.gb10.hide()

        vbox_main = QVBoxLayout()
        vbox_main.addStretch()
        vbox_main.addWidget(self.gb1)
        vbox_main.addWidget(self.gb2)
        vbox_main.addWidget(self.gb3)
        vbox_main.addWidget(self.gb4)
        vbox_main.addWidget(self.gb5)
        vbox_main.addWidget(self.gb6)
        vbox_main.addWidget(self.gb7)
        vbox_main.addWidget(self.gb8)
        vbox_main.addWidget(self.gb9)
        vbox_main.addWidget(self.gb10)
        vbox_main.addStretch()
        
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.methode_label)
        hbox2.addWidget(self.methode)
        hbox2.addWidget(self.barre)
        hbox2.addWidget(self.gaus_btn)
        hbox2.addWidget(self.rock_check)
        hbox2.addWidget(self.mov_avg_btn)
        hbox2.addWidget(self.mov_avg_val)
        hbox2.addWidget(self.lin_btn)
        hbox2.addStretch()
        hbox2.addWidget(self.img_lbl)
        vbox.addLayout(hbox2)

        hbox_main = QHBoxLayout()
        hbox_main.addLayout(vbox_main)
        vbox_can = QVBoxLayout()
        
        
        self.figure  = Figure()
        self.axes    = self.figure.add_subplot(111)
        self.figure.subplots_adjust(left=0.1,right=0.98,bottom=0.1,top=0.95)
        self.canvas  = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)

        vbox_can.addWidget(self.canvas)

        hbox = QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(self.toolbar)
        hbox.addStretch()
        vbox_can.addLayout(hbox)
        hbox_main.addLayout(vbox_can)
        vbox.addLayout(hbox_main)

    def ClearAxes(self):
        self.axes.clear()
        self.canvas.draw()

    def Plot(self):
        
        self.axes.set_xlabel('Angle [deg]')
        self.axes.set_ylabel('Intensity [a.u.]')
        self.axes.set_title('Data Graph')
        self.axes.plot(self.x,self.y, label='data', color = 'black')
        self.axes.legend()
        self.canvas.draw()
        self.gaus_btn.setEnabled(True)
        self.mov_avg_val.setEnabled(True)
        self.lin_btn.setEnabled(True)

    def open_file(self) :
        self.x = []
        self.y = []
        openDir = './'
        fname = QFileDialog.getOpenFileName(self, 'Choose the file containing the data', openDir, 'XRD data file (*.dql)')
        self.file_path = fname[0]
        if self.file_path == "" : return
        try :
            with open(self.file_path) as data :
                first = data.readline()
                while not("Angle" in first) :
                    first = data.readline()
                a = data.readlines()
                for elt in a :
                    b = elt.split(",")
                    self.x.append(float(b[0]))
                    self.y.append(float(b[1]))
            if self.mw.clplt : self.ClearAxes()
            self.hide_info_boxes()
            self.Plot()
        except :
            QMessageBox.information(None, 'Information', 'It seems that the format of the file you chose is not supported.', QMessageBox.Ok)

    def open_text_file(self) :
        self.x = []
        self.y = []
        openDir = './'
        fname = QFileDialog.getOpenFileName(self, 'Choose the file containing the data', openDir, 'text data file (*.txt)')
        self.file_path = fname[0]
        if self.file_path == "" : return
        self.text_file_prop()
        if (self.line_num == "" or self.sepa == "" or self.line_num == None or self.sepa == None) : return
        try :
            with open(self.file_path) as data :
                first = data.readline()
                for i in range(self.line_num-2) :
                    first = data.readline()
                a = data.readlines()
                for elt in a :
                    b = elt.split(self.sepa)
                    self.x.append(float(b[0]))
                    self.y.append(float(b[1]))
            if self.mw.clplt : self.ClearAxes()
            self.hide_info_boxes()
            self.Plot()
        except :
            QMessageBox.information(None, 'Information', 'It seems that the format of the file you chose is not supported or the file properties you specified are incorrect.', QMessageBox.Ok)

    def open_excel_file(self) :
        self.x = []
        self.y = []
        self.check = True
        openDir = './'
        fname = QFileDialog.getOpenFileName(self, 'Choose the file containing the data', openDir, 'excel data file (*.xlsx *.xlsm *.xls *.xlt *.xltm *.xlts *.xlsb)')
        self.file_path = fname[0]
        if self.file_path == "" : return
        self.excel_file_prop()
        if (self.excel_line_num == None or self.excel_line_num == "" or self.excel_column_num == "" or self.excel_column_num == None or (self.check and (self.excel_sheet_name == "" or self.excel_sheet_name == None))) : return
        try :
            wb = op.load_workbook (self.file_path)          
            for ws in wb. worksheets :                      
                Lfin = ws.max_row
                if self.check :
                    if (ws.title == self.excel_sheet_name) :
                        for row in ws.rows[self.excel_line_num-1: Lfin] :
                            self.x.append(row[self.excel_column_num-1].value)
                            self.y.append(row[self.excel_column_num].value)
                else :
                    for row in ws.rows[self.excel_line_num-1: Lfin] :
                        self.x.append(row[self.excel_column_num-1].value)
                        self.y.append(row[self.excel_column_num].value)
            if self.mw.clplt : self.ClearAxes()
            self.hide_info_boxes()
            self.Plot()
        except :
            QMessageBox.information(None, 'Information', 'It seems that the format of the file you chose is not supported or the file properties you specified are incorrect.', QMessageBox.Ok)

    def excel_file_prop(self) :
        Q = QMessageBox(self)
        Q.setWindowTitle("File properties")
        self.okbouton = QPushButton("Ok")
        self.okbouton.minimumSizeHint()
        self.okbouton.setEnabled(False)
        q = Q.layout()
        self.okbouton.clicked.connect(Q.close)
            
        hlayout1=QHBoxLayout(self)
        hlayout1.addStretch()
        hlayout1.addWidget(self.okbouton)
        hlayout1.addStretch()

        q.itemAtPosition(q.rowCount()-1,0).widget().hide()
        p0 = QLabel('Name of the sheet containing data')
        p1 = QLabel('N° of the first line containing data')
        p2 = QLabel('N° of the first column conaining data')
        self.L1 = QLineEdit()
        self.L2 = QLineEdit()
        self.L0 = QLineEdit()
        self.c = QCheckBox('Only one sheet')
        self.c.setChecked(False)
        self.c.stateChanged.connect(self.num_sheet)
        self.L1.textEdited[str].connect(self.excel_line_numb)
        self.L2.textEdited[str].connect(self.excel_column_numb)
        self.L0.textEdited[str].connect(self.excel_sheet_name_set)
        q.addWidget(p0,0,0)
        q.addWidget(p1,1,0)
        q.addWidget(p2,2,0)
        q.addWidget(self.L0,0,1,1,2)
        q.addWidget(self.c,0,3)
        q.addWidget(self.L1,1,1,1,2)
        q.addWidget(self.L2,2,1,1,2)
        q.addLayout(hlayout1,3,0,1,3)
        Q.exec()
        
    def text_file_prop(self) :
        Q = QMessageBox(self)
        Q.setWindowTitle("File properties")
        self.okbtn = QPushButton("Ok")
        self.okbtn.minimumSizeHint()
        self.okbtn.setEnabled(False)
        q = Q.layout()
        self.okbtn.clicked.connect(Q.close)
            
        hlayout1=QHBoxLayout(self)
        hlayout1.addStretch()
        hlayout1.addWidget(self.okbtn)
        hlayout1.addStretch()

        q.itemAtPosition(q.rowCount()-1,0).widget().hide()
        p1 = QLabel('N° of the first line containing data')
        p2 = QLabel('''Separator between x and y data (ex : ',' or ';' or '-')''')
        self.l1 = QLineEdit()
        self.l2 = QLineEdit()
        self.l1.textEdited[str].connect(self.line_numb)
        self.l2.textChanged[str].connect(self.separator)
        q.addWidget(p1,0,0)
        q.addWidget(p2,1,0)
        q.addWidget(self.l1,0,1,1,2)
        q.addWidget(self.l2,1,1,1,2)
        q.addLayout(hlayout1,2,0,1,4)
        Q.exec()

    def line_numb(self, text) :
        a = self.l1.cursorPosition()
        if len(text) == 0 :
            self.okbtn.setEnabled(False)
            return
        elif not(text[a-1] in "0123456789") :
            QToolTip.showText(QDesktopWidget().availableGeometry().center() + QPoint(100,-39),"Only numbers are allowed",self.l1)
            if len(text) == 1 :
                self.l1.setText("")
                self.line_num = None
                self.okbtn.setEnabled(False)
            else :
                self.l1.setText(text[:a-1]+text[a:])
                self.l1.setCursorPosition(a-1)
                self.line_num = int(text[:a-1]+text[a:])
                if (self.sepa != "" and self.sepa != None) : self.okbtn.setEnabled(True)
        else :
            self.line_num = int(text)
            if (self.sepa != "" and self.sepa != None) : self.okbtn.setEnabled(True)

    def excel_line_numb(self,text) :
        a = self.L1.cursorPosition()
        if len(text) == 0 :
            self.okbouton.setEnabled(False)
        elif not(text[a-1] in "0123456789") :
            QToolTip.showText(QDesktopWidget().availableGeometry().center() + QPoint(100,-39),"Only numbers are allowed",self.L1)
            if len(text) == 1 :
                self.L1.setText("")
                self.excel_line_num = None
                self.okbouton.setEnabled(False)
            else :
                self.L1.setText(text[:a-1]+text[a:])
                self.L1.setCursorPosition(a-1)
                self.excel_line_num = int(text[:a-1]+text[a:])
                if (self.excel_column_num != "" and self.excel_column_num != None and (not(self.check) or (self.excel_sheet_name != "" and self.excel_sheet_name != None))) : self.okbouton.setEnabled(True)
        else :
            self.excel_line_num = int(text)
            if (self.excel_column_num != "" and self.excel_column_num != None and (not(self.check) or (self.excel_sheet_name != "" and self.excel_sheet_name != None))) : self.okbouton.setEnabled(True)

    def excel_column_numb(self,text) :
        a = self.L2.cursorPosition()
        if len(text) == 0 :
            self.okbouton.setEnabled(False)
        elif not(text[a-1] in "0123456789") :
            QToolTip.showText(QDesktopWidget().availableGeometry().center() + QPoint(100,-39),"Only numbers are allowed",self.L2)
            if len(text) == 1 :
                self.L2.setText("")
                self.excel_column_num = None
                self.okbouton.setEnabled(False)
            else :
                self.L2.setText(text[:a-1]+text[a:])
                self.L2.setCursorPosition(a-1)
                self.excel_column_num = int(text[:a-1]+text[a:])
                if (self.excel_line_num != "" and self.excel_line_num != None and (not(self.check) or (self.excel_sheet_name != "" and self.excel_sheet_name != None))) : self.okbouton.setEnabled(True)
        else :
            self.excel_column_num = int(text)
            if (self.excel_line_num != "" and self.excel_line_num != None and (not(self.check) or (self.excel_sheet_name != "" and self.excel_sheet_name != None))) : self.okbouton.setEnabled(True)

    def excel_sheet_name_set(self,text) :
        if (self.excel_line_num != "" and self.excel_line_num != None and self.excel_column_num != "" and self.excel_column_num != None) : self.okbouton.setEnabled(True)
        if len(text) == 0 : self.okbouton.setEnabled(False)
        self.excel_sheet_name = str(text)

    def num_sheet(self) :
        if self.check :
            self.L0.setEnabled(False)
            self.check = False
        else :
            self.L0.setEnabled(True)
            self.check = True
            if (self.excel_column_num != "" and self.excel_column_num != None and self.excel_line_num != "" and self.excel_line_num != None and self.excel_sheet_name != "" and self.excel_sheet_name != None) : self.okbouton.setEnabled(True)
        return None
            
    def separator(self, text) :
        if len(text) == 0 :
            self.okbtn.setEnabled(False)
        self.sepa = str(text)
        if (self.sepa != "" and self.sepa != None) :
            if (self.line_num != "" and self.line_num != None) : self.okbtn.setEnabled(True)

    def gaussian_fit(self) :
        if (len(self.x) == 0 or len(self.y) == 0) :
            QMessageBox.information(None, 'Information', 'It seems that there is an issue with the data. Please check your file.', QMessageBox.Ok)
            return
        self.pic_num = 0
        r = False
        self.hide_info_boxes()
        if self.test_rock :
            self.threshold = 0
            param = [[0,len(self.x)-1,self.x[self.y.index(max(self.y))],max(self.y)]]
        else :
            self.find_threshold()
            try : param = self.find_pics()
            except :
                QMessageBox.information(None, 'Information', 'Unable to find a peak to analyse.', QMessageBox.Ok)
                return
        print(self.threshold)
        def gaus(X,amp,cent,sigma,d,e) :
            return amp*np.exp(-(X-cent)**2/(2*sigma**2))+d*X+e
        self.ClearAxes()
        self.Plot()
        for beg,end,center,amplitude in param :
            x = np.array(self.x[beg:end])
            y = np.array(self.y[beg:end])
            print(x[0])
            print(x[-1])
            try :
                popt,pcov = curve_fit(gaus,x,y,p0=[amplitude,center,1,1,1],maxfev=10000)
                am = gaus(popt[1],popt[0],popt[1],popt[2],popt[3],popt[4])
                print('am = ',am)
                if (am > 0.7*max(self.y[beg:end]) and am < 1.3*max(self.y[beg:end])) :
                    print('Pas overlap')
                    self.axes.plot(x,gaus(x,*popt), label='Peak n°'+str(self.pic_num+1)+' fit')
                    self.axes.legend()
                    self.canvas.draw()
                    fwhm = self.find_fwhm(popt,beg,end)
                    self.pic_num += 1
                    self.show_info_boxes(False,round(am,0),round(popt[1],3),round(fwhm,3))
                else :
                    x,y,m = self.increase_coords(x,y,beg,end,500)
                    try :
                        popt,pcov = curve_fit(gaus,x,y,p0=[amplitude,center,1,1,1],maxfev=100000)
                        am = gaus(popt[1],popt[0],popt[1],popt[2],popt[3],popt[4])
                        print('am_augmenté = ',am)
                        if (am > 0.7*max(self.y[beg:end]) and am < 1.3*max(self.y[beg:end])) :
                            print('Fonctionne avec les coordonnées augmentées')
                            self.axes.plot(x[(m-10):(10-m)],gaus(x[(m-10):(10-m)],*popt), label='Peak n°'+str(self.pic_num+1)+' fit')
                            self.axes.legend()
                            self.canvas.draw()
                            fwhm = self.find_fwhm(popt,beg,end)
                            self.pic_num += 1
                            self.show_info_boxes(False,round(am,0),round(popt[1],3),round(fwhm,3))
                        else :
                            print('Ne fonctionne pas avec les coordonnées augmentées, peut-être Overlap ?')
                            self.overlap(beg,end,amplitude)
                    except : print('ERROR avec coord augmentées')
                if ((fwhm == 0 or fwhm == None) and  not(self.test_rock)) :
                    print('Overlap')
                    self.overlap(beg,end,amplitude)
                elif (fwhm == 0 or fwhm == None) : QMessageBox.information(None, 'Information', 'Unable to find the coefficients for the gaussian fit on one or more peak(s).', QMessageBox.Ok)
            except RuntimeError :
                x,y,m = self.increase_coords(x,y,beg,end,500)
                try :
                    popt,pcov = curve_fit(gaus,x,y,p0=[amplitude,center,1,1,1],maxfev=100000)
                    am = gaus(popt[1],popt[0],popt[1],popt[2],popt[3],popt[4])
                    print('am_augmenté = ',am)
                    if (am > 0.7*max(self.y[beg:end]) and am < 1.3*max(self.y[beg:end])) :
                        print('Fonctionne avec les coordonnées augmentées')
                        self.axes.plot(x[(m-10):(10-m)],gaus(x[(m-10):(10-m)],*popt), label='Peak n°'+str(self.pic_num+1)+' fit')
                        self.axes.legend()
                        self.canvas.draw()
                        fwhm = self.find_fwhm(popt,beg,end)
                        self.pic_num += 1
                        self.show_info_boxes(False,round(am,0),round(popt[1],3),round(fwhm,3))
                    else :
                        print('Ne fonctionne pas avec les coordonnées augmentées')
                        y = self.moving_avrg_gaus(y,1)
                        try :
                            popt,pcov = curve_fit(gaus,x,y,p0=[amplitude,center,1,1,1],maxfev=100000)
                            am = gaus(popt[1],popt[0],popt[1],popt[2],popt[3],popt[4])
                            print('am_lissé = ', am)
                            if (am > 0.7*max(self.y[beg:end]) and am < 1.3*max(self.y[beg:end])) :
                                print('Fonctionne avec lissage')
                                self.axes.plot(x,gaus(x,*popt), label='Peak n°'+str(self.pic_num+1)+' fit')
                                self.axes.legend()
                                self.canvas.draw()
                                fwhm = self.find_fwhm(popt,beg,end)
                                self.pic_num += 1
                                self.show_info_boxes(False,round(am,0),round(popt[1],3),round(fwhm,3))
                            else :
                                self.pic_num += 1
                                print('Ne fonctionne avec rien')
                                r = True
                        except : print('ERROR avec lissage')
                except : print('ERROR avec coord augmentées')
            except :
                r = True
                print('ERROR')
        if r : QMessageBox.information(None, 'Information', 'Unable to find the coefficients for the gaussian fit on one or more peak(s).', QMessageBox.Ok)

    def overlap(self,b,e,a) :
        def gaus(X,amp,cent,sigma,d,e) :
            return amp*np.exp(-(X-cent)**2/(2*sigma**2))+d*X+e
        r = False
        mid_1 = int((b+e)/2)
        max_1, max_2 = self.y.index(max(self.y[b:mid_1])), self.y.index(max(self.y[mid_1:e]))
        mid = self.y.index(min(self.y[max_1:max_2]))
        begin, ending = 0,0
        for i,eltx in enumerate(self.y[b:mid]) :
            if eltx>=self.y[mid] and begin == 0 :
                begin = i
        for j,elty in enumerate(self.y[mid+10:e]) :
            if elty<=self.y[mid] and ending == 0:
                ending = j
        x1, x2 = np.array(self.x[begin+b:mid]), np.array(self.x[mid:mid+11+ending])
        y1, y2 = np.array(self.y[begin+b:mid]), np.array(self.y[mid:mid+11+ending])
        popt1, pcov1 = curve_fit(gaus,x1,y1,p0=[max(y1),self.x[int((b+begin+mid)/2)],1,1,1],maxfev=10000)
        popt2, pocv2 = curve_fit(gaus,x2,y2,p0=[max(y2),self.x[int((mid+mid+11+ending)/2)],1,1,1],maxfev=10000)
        am1 = gaus(popt1[1],popt1[0],popt1[1],popt1[2],popt1[3],popt1[4])
        am2 = gaus(popt2[1],popt2[0],popt2[1],popt2[2],popt2[3],popt2[4])
        print(am1)
        if (am1 > 0.7*max(y1) and am1 < 1.3*max(y1)) :
            self.axes.plot(x1,gaus(x1,*popt1), label='Peak n°'+str(self.pic_num+1)+' fit')
            self.axes.legend()
            self.canvas.draw()
            fwhm = round(self.find_fwhm(popt1,begin+b,mid),3)
            self.pic_num += 1
            self.show_info_boxes(False,round(am1,0),round(popt1[1],3),fwhm)
        else :
            x1, y1, m = self.increase_coords(x1,y1,begin+b,mid,1000)
            popt1, pcov1 = curve_fit(gaus,x1,y1,p0=[max(y1),self.x[int((b+begin+mid)/2)],1,1,1],maxfev=10000)
            am1 = gaus(popt1[1],popt1[0],popt1[1],popt1[2],popt1[3],popt1[4])
            print(am1)
            if (am1 > 0.7*max(y1) and am1 < 1.3*max(y1)) :
                self.axes.plot(x1[(m-10):(10-m)],gaus(x1[(m-10):(10-m)],*popt1), label='Peak n°'+str(self.pic_num+1)+' fit')
                self.axes.legend()
                self.canvas.draw()
                fwhm = round(self.find_fwhm(popt1,begin+b,mid),3)
                self.pic_num += 1
                self.show_info_boxes(False,round(am1,0),round(popt1[1],3),fwhm)
            else :
                if r :
                    r = False
                else :
                    r = True
                self.pic_num += 1
        if (am2 > 0.7*max(y2) and am2 < 1.3*max(y2)) :
            self.axes.plot(x2,gaus(x2,*popt2), label='Peak n°'+str(self.pic_num+1)+' fit')
            self.axes.legend()
            self.canvas.draw()
            fwhm = round(self.find_fwhm(popt2,mid,mid+11+ending),3)
            self.pic_num += 1
            self.show_info_boxes(False,round(am2,0),round(popt2[1],3),fwhm)
        else :
            x2, y2, m = self.increase_coords(x2,y2,mid,mid+11+ending,1000)
            popt2, pocv2 = curve_fit(gaus,x2,y2,p0=[max(y2),self.x[int((mid+mid+11+ending)/2)],1,1,1],maxfev=10000)
            am2 = gaus(popt2[1],popt2[0],popt2[1],popt2[2],popt2[3],popt2[4])
            if (am2 > 0.7*max(y2) and am2 < 1.3*max(y2)) :
                self.axes.plot(x2[(m-10):(10-m)],gaus(x2[(m-10):(10-m)],*popt2), label='Peak n°'+str(self.pic_num+1)+' fit')
                self.axes.legend()
                self.canvas.draw()
                fwhm = round(self.find_fwhm(popt2,mid,mid+11+ending),3)
                self.pic_num += 1
                self.show_info_boxes(False,round(am2,0),round(popt2[1],3),fwhm)
            else :
                r = True
                self.pic_num += 1
        if r : QMessageBox.information(None, 'Information', 'Unable to find the coefficients for the gaussian fit on one or more peak(s).', QMessageBox.Ok)

    def find_threshold(self) :
        y = np.copy(self.y)
        y.sort()
        self.threshold = y[int(len(y)/5)] + 10
        print("Threshold = ",self.threshold)

    def rocking_curve_active(self) :
        if self.test_rock :
            self.test_rock = False
        else :
            self.test_rock = True

    def hide_info_boxes(self) :
        self.gb1.hide()
        self.gb2.hide()
        self.gb3.hide()
        self.gb4.hide()
        self.gb5.hide()
        self.gb6.hide()
        self.gb7.hide()
        self.gb8.hide()
        self.gb9.hide()
        self.gb10.hide()

    def show_info_boxes(self,err,a,c,f) :
        if self.pic_num == 1 :
            self.gb1.show()
            if err :
                self.amp_1.setText('Amplitude : Error')
                self.center_1.setText('Center : Error')
                self.fwhm_1.setText('FWHM : Error')
            else :
                self.amp_1.setText('Amplitude : '+str(a))
                self.center_1.setText('Center : '+str(c))
                self.fwhm_1.setText('FWHM : '+str(f))
        if self.pic_num == 2 :
            self.gb2.show()
            if err :
                self.amp_2.setText('Amplitude : Error')
                self.center_2.setText('Center : Error')
                self.fwhm_2.setText('FWHM : Error')
            else :
                self.amp_2.setText('Amplitude : '+str(a))
                self.center_2.setText('Center : '+str(c))
                self.fwhm_2.setText('FWHM : '+str(f))
        if self.pic_num == 3 :
            self.gb3.show()
            if err :
                self.amp_3.setText('Amplitude : Error')
                self.center_3.setText('Center : Error')
                self.fwhm_3.setText('FWHM : Error')
            else :
                self.amp_3.setText('Amplitude : '+str(a))
                self.center_3.setText('Center : '+str(c))
                self.fwhm_3.setText('FWHM : '+str(f))
        if self.pic_num == 4 :
            self.gb4.show()
            if err :
                self.amp_4.setText('Amplitude : Error')
                self.center_4.setText('Center : Error')
                self.fwhm_4.setText('FWHM : Error')
            else :
                self.amp_4.setText('Amplitude : '+str(a))
                self.center_4.setText('Center : '+str(c))
                self.fwhm_4.setText('FWHM : '+str(f))
        if self.pic_num == 5 :
            self.gb5.show()
            if err :
                self.amp_5.setText('Amplitude : Error')
                self.center_5.setText('Center : Error')
                self.fwhm_5.setText('FWHM : Error')
            else :
                self.amp_5.setText('Amplitude : '+str(a))
                self.center_5.setText('Center : '+str(c))
                self.fwhm_5.setText('FWHM : '+str(f))
        if self.pic_num == 6 :
            self.gb6.show()
            if err :
                self.amp_6.setText('Amplitude : Error')
                self.center_6.setText('Center : Error')
                self.fwhm_6.setText('FWHM : Error')
            else :
                self.amp_6.setText('Amplitude : '+str(a))
                self.center_6.setText('Center : '+str(c))
                self.fwhm_6.setText('FWHM : '+str(f))
        if self.pic_num == 7 :
            self.gb7.show()
            if err :
                self.amp_7.setText('Amplitude : Error')
                self.center_7.setText('Center : Error')
                self.fwhm_7.setText('FWHM : Error')
            else :
                self.amp_7.setText('Amplitude : '+str(a))
                self.center_7.setText('Center : '+str(c))
                self.fwhm_7.setText('FWHM : '+str(f))
        if self.pic_num == 8 :
            self.gb8.show()
            if err :
                self.amp_8.setText('Amplitude : Error')
                self.center_8.setText('Center : Error')
                self.fwhm_8.setText('FWHM : Error')
            else :
                self.amp_8.setText('Amplitude : '+str(a))
                self.center_8.setText('Center : '+str(c))
                self.fwhm_8.setText('FWHM : '+str(f))
        if self.pic_num == 9 :
            self.gb9.show()
            if err :
                self.amp_9.setText('Amplitude : Error')
                self.center_9.setText('Center : Error')
                self.fwhm_9.setText('FWHM : Error')
            else :
                self.amp_9.setText('Amplitude : '+str(a))
                self.center_9.setText('Center : '+str(c))
                self.fwhm_9.setText('FWHM : '+str(f))
        if self.pic_num == 10 :
            self.gb10.show()
            if err :
                self.amp_10.setText('Amplitude : Error')
                self.center_10.setText('Center : Error')
                self.fwhm_10.setText('FWHM : Error')
            else :
                self.amp_10.setText('Amplitude : '+str(a))
                self.center_10.setText('Center : '+str(c))
                self.fwhm_10.setText('FWHM : '+str(f))
        
    
    def find_fwhm(self,param,b,e) :
        def gaus(X,amp,cent,sigma,d,e) :
            return amp*np.exp(-(X-cent)**2/(2*sigma**2))+d*X+e
        x = np.arange(self.x[b],self.x[e]+0.001,0.001)
        y = gaus(x,*param)
        y = y.tolist()
        maxi = param[0]
        target = maxi/2
        init = y.index(max(y))
        pre_mid, post_mid, pre_pos, post_pos = -np.inf, np.inf, 0, 0
        for i,pre in enumerate(y[:init]) :
            if pre > target and pre_mid <= target :
                pre_mid = pre
                pre_pos = i
        for j,post in enumerate(y[init:]) :
            if post < target and post_mid >= target :
                post_mid = post
                post_pos = j
        return abs(x[init+post_pos]-x[pre_pos])
                
    def find_pics(self) :
        old = self.y[0]
        pos = []
        center = 0
        amplitude = 0
        for i,elt in enumerate(self.y) :
            if old <= self.threshold and elt > self.threshold :
                a = [i]
            if old > self.threshold and elt <= self.threshold :
                if i-a[0] >= 60 :
                    a.append(i)
                    pos.append(a)
            old = elt
        for i,elt in enumerate(pos) :
            width = pos[i][1] - pos[i][0]
            pos[i][0]=int(pos[i][0]//10*10-(0.25*width))
            pos[i][1]=int(pos[i][1]//10*10+(0.25*width))
            center = round(self.x[pos[i][0]+self.y[pos[i][0]:pos[i][1]].index(max(self.y[pos[i][0]:pos[i][1]]))],3)
            print("beg = ", self.x[pos[i][0]])
            print("end = ", self.x[pos[i][1]])
            print("center = ", center)
            amplitude = int(max(self.y[elt[0]:elt[1]]))
            pos[i].append(center)
            pos[i].append(amplitude)
        print('pos : ',pos)
        return pos    

    def increase_coords(self,x,y,b,e,n) :
        print("a")
        borne_inf, borne_sup = np.inf, np.inf
        print("b")
        if b-n < 0 :
            print("c")
            borne_inf = b
            print("d")
        if e+n > len(self.x) :
            print("e")
            borne_sup = len(self.x)-e
            print("f")
        print("g")
        m = min(borne_inf, borne_sup, n)
        print("h")
        return (np.array(self.x[b-m:e+m]), np.hstack([min(y[0],y[-1])*np.ones(m),y,min(y[0],y[-1])*np.ones(m)]),m)
    
    def moving_avrg_gaus(self,y,val) :
        y_avg = np.copy(y)
        for i in range(val,len(y_avg)-val) :
            y_avg[i] = np.mean(y[i-val:i+val])
        return y_avg

    def moving_avrg(self) :
        if (len(self.x) == 0 or len(self.y) == 0) :
            QMessageBox.information(None, 'Information', 'It seems that there is an issue with the data. Please check your file.', QMessageBox.Ok)
            return
        self.hide_info_boxes()
        val = self.mov_avg_num
        y_avg = np.copy(self.y)
        for i in range(val,len(y_avg)-val) :
            y_avg[i] = np.mean(self.y[i-val:i+val])
        self.ClearAxes()
        self.Plot()
        self.axes.plot(self.x,y_avg, label='Moving average')
        self.axes.legend()
        self.canvas.draw()

    def mov_avg_value(self,text) :
        a = self.mov_avg_val.cursorPosition()
        if len(text) == 0 :
            self.mov_avg_btn.setEnabled(False)
        elif not(text[a-1] in "0123456789") :
            QToolTip.showText(QDesktopWidget().availableGeometry().center() + QPoint(100,-39),"Only numbers are allowed",self.mov_avg_val)
            if len(text) == 1 :
                self.mov_avg_val.setText("")
                self.mov_avg_num = None
                self.mov_avg_btn.setEnabled(False)
            else :
                self.mov_avg_val.setText(text[:a-1]+text[a:])
                self.mov_avg_val.setCursorPosition(a-1)
                self.mov_avg_num = int(text[:a-1]+text[a:])
                self.mov_avg_btn.setEnabled(True)
        else :
            self.mov_avg_num = int(text)
            self.mov_avg_btn.setEnabled(True)

    def linear_fit(self) :
        if (len(self.x) == 0 or len(self.y) == 0) :
            QMessageBox.information(None, 'Information', 'It seems that there is an issue with the data. Please check your file.', QMessageBox.Ok)
            return
        self.hide_info_boxes()
        x, y = np.array(self.x), np.array(self.y)
        def linear(X, a, b) :
            return a*X+b
        try :
            popt, pcov = curve_fit(linear,x,y)
            self.ClearAxes()
            self.Plot()
            self.axes.plot(x,linear(x,*popt),label='Linear fit (y='+str(round(popt[0],3))+'.x'+str((popt[1]>0)*'+')+str((popt[1]<0)*'-')+str(abs(round(popt[1],3)))+')')
            self.axes.legend()
            self.canvas.draw()
        except :
            QMessageBox.information(None, 'Information', 'Unable to find the coefficients for the linear fit', QMessageBox.Ok)

    def combobox(self) :
        a = self.methode.currentIndex()
        if a == 0 :
            self.gaus_btn.hide()
            self.mov_avg_btn.hide()
            self.mov_avg_val.hide()
            self.barre.hide()
            self.rock_check.hide()
            self.lin_btn.hide()
            if self.x != None :
                if len(self.x) != 0 :
                    self.ClearAxes()
                    self.Plot()
        elif a == 1 :
            self.mov_avg_btn.hide()
            self.mov_avg_val.hide()
            self.barre.show()
            self.gaus_btn.show()
            self.rock_check.show()
            self.lin_btn.hide()
        elif a == 2 :
            self.gaus_btn.hide()
            self.rock_check.hide()
            self.mov_avg_btn.show()
            self.mov_avg_val.show()
            self.barre.show()
            self.lin_btn.hide()
        elif a == 3 :
            self.gaus_btn.hide()
            self.rock_check.hide()
            self.mov_avg_btn.hide()
            self.mov_avg_val.hide()
            self.barre.hide()
            self.lin_btn.show()

if __name__ == '__main__' :
    app    = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path('./icones/app_logo.png')))
    my_app = DataAnalyser()
    return_status = app.exec_() 
    sys.exit(return_status)
